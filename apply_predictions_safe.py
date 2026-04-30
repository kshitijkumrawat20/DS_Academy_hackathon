import os
import pandas as pd
import numpy as np
from xgboost import XGBClassifier
from sklearn.preprocessing import LabelEncoder
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# 1. Load 2021 Data to Train the Model
data_dir = 'data'
files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx') and not f.startswith('~$') and 'Template' not in f]

dfs = []
for file in files:
    file_path = os.path.join(data_dir, file)
    print(f"Reading training data from: {file}...")
    df_temp = pd.read_excel(file_path, header=None)
    try:
        header_idx = df_temp[df_temp[0] == 'STATE/UT NAME'].index[0]
    except:
        header_idx = 2
        
    df = pd.read_excel(file_path, header=header_idx)
    df = df.rename(columns={
        'STATE/UT NAME': 'State',
        'AC NO.': 'Constituency_No',
        'AC NAME': 'Constituency_Name',
        'CANDIDATE NAME': 'Candidate Name',
        'PARTY': 'Party',
        'TOTAL': 'Total_Votes'
    })
    
    if 'Constituency_Name' not in df.columns: continue
    
    df = df.dropna(subset=['Candidate Name', 'Constituency_Name', 'Total_Votes'])
    df = df[df['Candidate Name'] != 'NOTA']
    
    df['Total_Votes'] = pd.to_numeric(df['Total_Votes'], errors='coerce').fillna(0)
    df['Win'] = df.groupby('Constituency_No')['Total_Votes'].transform(lambda x: (x == x.max()).astype(int))
    
    df['State'] = df['State'].astype(str).str.title().str.strip()
    df = df[['State', 'Constituency_Name', 'Party', 'Win']]
    dfs.append(df)

train_df = pd.concat(dfs, ignore_index=True)

major_parties = ['BJP', 'INC', 'AITC', 'DMK', 'AIADMK', 'CPI(M)', 'JD(U)', 'AGP', 'AIUDF', 'TMC', 'IUML']
train_df['is_major_party'] = train_df['Party'].apply(lambda x: 1 if any(p in str(x).upper() for p in major_parties) else 0)

le_party = LabelEncoder()
le_const = LabelEncoder()
le_state = LabelEncoder()

class SafeEncoder:
    def __init__(self, le):
        self.le = le
    def fit(self, data):
        self.le.fit(list(data) + ['UNKNOWN'])
    def transform(self, data):
        data = [x if x in self.le.classes_ else 'UNKNOWN' for x in data]
        return self.le.transform(data)

safe_party = SafeEncoder(le_party)
safe_const = SafeEncoder(le_const)
safe_state = SafeEncoder(le_state)

template_file = 'data/Submission Template.xlsx'
sheets = ['Assam', 'Kerala', 'Puducherry', 'Tamil Nadu', 'West Bengal']

all_parties = set(train_df['Party'].dropna().astype(str))
all_consts = set(train_df['Constituency_Name'].dropna().astype(str))
all_states = set(train_df['State'].dropna().astype(str))

test_dict = {}
for sheet in sheets:
    test_df = pd.read_excel(template_file, sheet_name=sheet)
    test_dict[sheet] = test_df
    all_parties.update(test_df['Party'].dropna().astype(str))
    col_const = 'Constituency' if 'Constituency' in test_df.columns else 'Constituency_Name'
    all_consts.update(test_df[col_const].dropna().astype(str))
    col_state = 'State/UT' if 'State/UT' in test_df.columns else 'State'
    all_states.update(test_df[col_state].dropna().astype(str))

safe_party.fit(all_parties)
safe_const.fit(all_consts)
safe_state.fit(all_states)

train_df['Party_encoded'] = safe_party.transform(train_df['Party'].astype(str))
train_df['Constituency_Name_encoded'] = safe_const.transform(train_df['Constituency_Name'].astype(str))
train_df['State_encoded'] = safe_state.transform(train_df['State'].astype(str))

features = ['Party_encoded', 'Constituency_Name_encoded', 'State_encoded', 'is_major_party']
X_train = train_df[features]
y_train = train_df['Win']

print("\nTraining XGBoost on 2021 ECI Data...")
model = XGBClassifier(n_estimators=150, learning_rate=0.08, max_depth=5, subsample=0.8, eval_metric='logloss', random_state=42)
model.fit(X_train, y_train)

print("\nEditing exactly Column F of the official template using openpyxl...")
wb = openpyxl.load_workbook(template_file)

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    
    test_df = test_dict[sheet_name].copy()
    test_df['is_major_party'] = test_df['Party'].apply(lambda x: 1 if any(p in str(x).upper() for p in major_parties) else 0)
    
    col_const = 'Constituency' if 'Constituency' in test_df.columns else 'Constituency_Name'
    col_state = 'State/UT' if 'State/UT' in test_df.columns else 'State'
    
    X_test = pd.DataFrame()
    X_test['Party_encoded'] = safe_party.transform(test_df['Party'].astype(str))
    X_test['Constituency_Name_encoded'] = safe_const.transform(test_df[col_const].astype(str))
    X_test['State_encoded'] = safe_state.transform(test_df[col_state].astype(str))
    X_test['is_major_party'] = test_df['is_major_party']
    
    # Predict Probabilities
    test_df['WinTarget_Prob'] = model.predict_proba(X_test)[:, 1]
    
    test_df['Predicted Outcome (W/L/O)'] = 'L'
    # Group by Constituency and pick max prob to assign the single W
    winners_idx = test_df.groupby(col_const)['WinTarget_Prob'].idxmax()
    test_df.loc[winners_idx, 'Predicted Outcome (W/L/O)'] = 'W'
    
    # Find Column F dynamically or fallback to 6
    pred_col_idx = None
    header_row = 1
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and 'Predicted Outcome (W/L/O)' in row:
            pred_col_idx = row.index('Predicted Outcome (W/L/O)') + 1
            header_row = row_idx
            break
            
    if pred_col_idx is None:
        pred_col_idx = 6 # Fallback to Column F
    
    # Write W/L directly into the opened Excel
    for i, pred in enumerate(test_df['Predicted Outcome (W/L/O)'], start=header_row + 1):
        ws.cell(row=i, column=pred_col_idx, value=pred)
        
    print(f"Updated exactly Column F for sheet: {sheet_name}")

output_file = 'IP2026_YOUR-EMAIL@DOMAIN.COM_YOUR-FULL-NAME.xlsx'
wb.save(output_file)
print(f"\nSUCCESS! File saved flawlessly enforcing all structural rules: {output_file}")
print("Please manually rename 'YOUR-EMAIL@DOMAIN.COM_YOUR-FULL-NAME' to your actual email and name before uploading.")